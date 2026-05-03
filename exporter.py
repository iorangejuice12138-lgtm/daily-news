"""Daily News Aggregator - Export Module (CSV & Excel)"""

import csv
import logging
import os
from datetime import datetime

import config

logger = logging.getLogger(__name__)

CSV_HEADERS = ["日期", "标签", "标题", "摘要", "来源", "链接"]


def _sort_by_tags(news_list: list[dict]) -> list[dict]:
    """Sort news items by primary tag, preserving tag order from config."""
    tag_order = {tag: i for i, tag in enumerate(config.TAGS)}
    tag_order["未分类"] = len(config.TAGS)

    def sort_key(item):
        primary_tag = item.get("tags", ["未分类"])[0]
        return tag_order.get(primary_tag, len(config.TAGS))

    return sorted(news_list, key=sort_key)


def _build_rows(news_list: list[dict], date_str: str) -> list[list[str]]:
    """Convert news items to table rows."""
    rows = []
    for item in news_list:
        tags_str = ", ".join(item.get("tags", []))
        rows.append([
            date_str,
            tags_str,
            item.get("title", ""),
            item.get("summary", ""),
            item.get("source", ""),
            item.get("url", ""),
        ])
    return rows


def export_csv(rows: list[list[str]]):
    """Append rows to the CSV file. Creates header if file is new."""
    os.makedirs(config.DATA_DIR, exist_ok=True)
    write_header = not os.path.exists(config.CSV_PATH) or os.path.getsize(config.CSV_PATH) == 0

    with open(config.CSV_PATH, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(CSV_HEADERS)
        writer.writerows(rows)

    logger.info(f"CSV: appended {len(rows)} rows to {config.CSV_PATH}")


def export_excel(rows: list[list[str]], date_str: str):
    """Append rows to the Excel file with formatting."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    os.makedirs(config.DATA_DIR, exist_ok=True)

    if os.path.exists(config.XLSX_PATH):
        wb = load_workbook(config.XLSX_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "每日新闻汇总"
        # Write header
        for col, header in enumerate(CSV_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        # Set column widths
        widths = [14, 16, 40, 50, 14, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Append rows
    for row_data in rows:
        ws.append(row_data)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(config.XLSX_PATH)
    logger.info(f"Excel: appended {len(rows)} rows to {config.XLSX_PATH}")


def export(news_list: list[dict], date_str: str):
    """Sort, build rows, and export to both CSV and Excel."""
    sorted_news = _sort_by_tags(news_list)
    rows = _build_rows(sorted_news, date_str)

    export_csv(rows)
    export_excel(rows, date_str)

    logger.info(f"Export complete: {len(rows)} news items for {date_str}")
